from os import listdir
from os.path import isfile, join
import os
import openpyxl
import string
import random

def main():
    curr_dir = os.path.dirname(os.path.realpath(__file__))

    print("Getting files in folder a")
    files_in_folder_a = [file for file in listdir(join(curr_dir, 'folder-a')) if isfile(join(curr_dir, 'folder-a', file))]
    
    print("Getting files in folder b")
    files_in_folder_b = [file for file in listdir(join(curr_dir, 'folder-b')) if isfile(join(curr_dir, 'folder-b', file))]

    print("Files in folder a", files_in_folder_a)

    print("Files in folder b", files_in_folder_b)

    dup_list = find_duplicate_file_in(folder_a=files_in_folder_a, folder_b=files_in_folder_b, 
        folder_a_path=join(curr_dir, 'folder-a'), folder_b_path=join(curr_dir, 'folder-b'))

    for dl in dup_list:
        merge_xlsx_files(dl)


def find_duplicate_file_in(folder_a, folder_b, folder_a_path, folder_b_path):

    intersected = set(folder_a).intersection(folder_b)
    print("dup files", intersected)

    # put dup files (array) in a list
    # /User/admin/folder-a/file-a.xlsx
    # /User/admin/folder-b/file-a.xlsx
    #[['/User/admin/folder-a/file-a.xlsx', '/User/admin/folder-b/file-a.xlsx'], [...]]

    dup_list = []
    for i in intersected:
        a = join(folder_a_path, i)
        b = join(folder_b_path, i)
        dup_list.append([a, b])

    print("dup list", dup_list)
    return dup_list


def merge_xlsx_files(file_paths):
    wb = openpyxl.Workbook()
    wb_name: str = get_random_string(5)
    index = 0

    for file_path in file_paths:
        wb_sheet = wb.active if index == 0 else wb.create_sheet()

        file_a_wb = openpyxl.load_workbook(join(file_path))
        sheet_a = file_a_wb.active    
        sheet_a_max_row = sheet_a.max_row
        sheet_a_max_col = sheet_a.max_column

        for row in range(1, sheet_a_max_row + 1):
            for cell in range(1, sheet_a_max_col + 1):
                wb_sheet.cell(row=row, column=cell).value = sheet_a.cell(row=row, column=cell).value

        print("xlsx", wb.active)
        curr_dir = os.path.dirname(os.path.realpath(__file__))
        output_file_path = join(curr_dir, 'output', (wb_name + '.xlsx'))
        wb.save(output_file_path)
        index = index + 1

    return True


def get_random_string(length):
    letters = string.ascii_lowercase
    result_str = ''.join(random.choice(letters) for i in range(length))
    print("Random string of length", length, "is:", result_str)
    return result_str


if __name__ == "__main__":
    main()
